from openpyxl import load_workbook
from openpyxl import Workbook
wb_w = Workbook()
dest_filename = 'Result.xlsx'
ws1 = wb_w.create_sheet(title='Results')
wb_val = load_workbook(filename='ExcelDocument.xlsx', data_only=True)
sheet_val = wb_val['Общая информация']
Percent = []
Pure_profit = []
FCF = []
def Finish_step():
    Pay_disc = Revenue_all_type()
    L_twenty = (-1*12)*10**6
    NPV = 0
    for i in range(len(Pay_disc)):
        NPV += Pay_disc[i]
    IRR = sheet_val['M32'].value
    PI = NPV/(-1*L_twenty)
    PI += 1
    PBP = []
    DPBP = []
    counter_pure = L_twenty
    counter_disc = L_twenty
    PBP_b = False
    Ages_PBP = 0
    DPBP_b = False
    Ages_DPBP = 0
    for i in range(5):
        counter_pure = counter_pure+FCF[i]
        counter_disc = counter_disc+Pay_disc[i]
        PBP.append(counter_pure)
        DPBP.append(counter_disc)
        if PBP[i]>0 and PBP_b == False:
            print('Need ', i+1, ' ', 'age(PBP)')
            PBP_b = True
            Ages_PBP = i+1
        if DPBP[i] > 0 and DPBP_b == False:
            print('Need ', i, ' ', 'age(DPBP)')
            DPBP_b = True
            Ages_DPBP = i
    print(PBP)
    print('IRR=', IRR, ' ', 'PI= ', PI, 'NPV= ', NPV, 'PBP= ', Ages_PBP, 'DPBP= ', Ages_DPBP)
    print(Pay_disc)
    print(PBP)
    print(DPBP)
    if (NPV >0 and PI >1 and IRR >Trebue_doxodn and Ages_PBP <5 and Ages_DPBP <5):
        ws1['A2'] = 'Инвестиционный проект является выгодным и соотствует всем условиям'
        ws1['A3'] = 'NPV = '+str(NPV)
        ws1['A5'] = "IRR = "+str(IRR)
        ws1['A7'] = 'PI = '+str(PI)
        ws1['A9'] = 'PBP = ' + str(Ages_PBP)+' '+'лет'
        ws1['A11'] = 'DBPB = ' + str(Ages_DPBP)+' '+'лет'
        wb_w.save(filename=dest_filename)
    else:
        ws1['A2'] = 'Инвестиционный проект не является выгодным, все параметры приведены ниже'
        ws1['A3'] = 'NPV = ' + str(NPV)
        ws1['A5'] = "IRR = " + str(IRR)
        ws1['A7'] = 'PI = ' + str(PI)
        ws1['A9'] = 'PBP = более 5-ти' + ' ' + 'лет'
        ws1['A11'] = 'DBPB = более 5-ти' + ' ' + 'лет'
        wb_w.save(filename=dest_filename)

def Revenue_all_type():
    Op_def = Operation_capital()
    Credit_func = Credit()
    Credit_func.append(0)
    Credit_func.append(0)
    Revenue_wo = []  # Выручка
    Revenue_wt = []  # Выручка минус налог
    Gross_profit = []  # Валовая прибыль
    Revenue_for_sales = []  # Прибыль от продаж
    Profit_before_tax = []
    Pure_profit_disc = []
    Percent.append(0)
    Percent.append(0)
    Pogashenie_dolg_kred[3] = 0
    Pogashenie_dolg_kred[4] = 0
    print("Pogashenie ", Pogashenie_dolg_kred)
    FCF.append(-1*12*10**6)
    for i in range(5):
        Revenue_wo.append(Max_production * Obyom_proizv[i] * Cost_per_unit)
        Revenue_wt.append(Revenue_wo[i] * (1 - Nalogovaya_nagruzka))
        Gross_profit.append(Revenue_wt[i] - (Max_production * Obyom_proizv[i] * Sebe_stoim)-Amortization)
        Revenue_for_sales.append(Gross_profit[i] - (Komm_rasx + Uprav_rasx)-Percent[i])
        Profit_before_tax.append(Revenue_for_sales[i]-Credit_func[i])
        if i == 4:
            Profit_before_tax[i]+=2*(10**6)
        Pure_profit.append(Profit_before_tax[i]*(1-Stavka_po_nalogu))
        FCF.append(Pure_profit[i]+Amortization-Op_def[i]-((8*10**6+Op_def[5])*Pogashenie_dolg_kred[i]))
        Pure_profit_disc.append(FCF[i+1]/(1+Trebue_doxodn)**(i+1))
    print('FCF - ', FCF)
    print('Дисконтированная прибыль - ', Pure_profit_disc)
    print(Pogashenie_dolg_kred)
    return Pure_profit_disc

def Credit():
    Dolg = []
    Payment = []
    Ostatok = []
    Op_def = (Operation_capital()[5])
    Credit_first_point = Stoimost_oborud - Akzion_finan_of_project
    Dolg.append((Credit_first_point+Op_def)*(1+Stavka_po_kreditu))
    Percent.append((Credit_first_point+Op_def)*Stavka_po_kreditu)
    Payment.append((Credit_first_point+Op_def)*Pogashenie_dolg_kred[0]+Percent[0])
    Ostatok.append(Dolg[0]-Payment[0])
    Ostatok.append((Credit_first_point+Op_def)*Pogashenie_dolg_kred[1])
    Percent.append(Ostatok[0]*Stavka_po_kreditu)
    Payment.append((Credit_first_point+Op_def)*Pogashenie_dolg_kred[1]+Percent[1])
    Dolg.append(Ostatok[0])
    Dolg.append(Ostatok[1])
    Percent.append(Ostatok[1]*Stavka_po_kreditu)
    Payment.append(Dolg[2]+Percent[2])
    print('Percent - ', Percent)
    return (Payment)

def Operation_capital():
    S_and_M_Per_Year = []
    NZP = []
    ZGP = []
    DZ = []
    KZ = []
    Diffe_value= []
    Summ=[]
    for i in range(5): #Заполнение массива с расчетами
        S_and_M_Per_Year.append(((Max_production*1000*Obyom_proizv[i]*S_and_M)/Number_of_part[0])//1000)
        NZP.append(Max_production*Obyom_proizv[i]*Sebe_stoim/Number_of_part[1])
        ZGP.append(Max_production*Obyom_proizv[i]*Sebe_stoim/Number_of_part[2])
        DZ.append((Vyruchka_po_godam[i]/Number_of_part[3])//1)
        KZ.append((Max_production*Obyom_proizv[i]*S_and_M/Number_of_part[4])//1)
    S_and_M_Per_Year.append(0)#Ввод последней строки с нулевыми значениями
    NZP.append(0)
    ZGP.append(0)
    DZ.append(0)
    KZ.append(0)
    for k in range(5):#Двумернй массив разности параметров
        Diffe_value.append([])
        for j in range(5):
            if j == 0:
                Diffe_value[k].append(S_and_M_Per_Year[k+1] - S_and_M_Per_Year[k])
            if j == 1:
                Diffe_value[k].append(NZP[k+1] - NZP[k])
            if j == 2:
                Diffe_value[k].append(ZGP[k+1] - ZGP[k])
            if j == 3:
                Diffe_value[k].append(DZ[k+1] - DZ[k])
            if j == 4:
                Diffe_value[k].append(KZ[k+1] - KZ[k])
    for k in range(5):
        Summ.append((Diffe_value[k][0]+Diffe_value[k][1]+Diffe_value[k][2]+Diffe_value[k][3]-Diffe_value[k][4]))
    Summ.append(S_and_M_Per_Year[0]+NZP[0]+ZGP[0]+DZ[0]-KZ[0])
    print('Sum ',Summ)
    return(Summ)

        # Общая информация

Stoimost_oborud = (sheet_val['B2'].value)*(10**6)
Srok_pol_ispolz = sheet_val['B3'].value
Max_production = sheet_val['B4'].value*1000
Akzion_finan_of_project = (sheet_val['B5'].value)*(10**6)
Trebue_doxodn = sheet_val['B6'].value
Stavka_po_kreditu = sheet_val['B7'].value
Stavka_po_nalogu = sheet_val['B8'].value
Nalogovaya_nagruzka = sheet_val['B9'].value
Amortization = 4*(10**6)
index = ['B', 'C', 'D', 'E', 'F']
Obyom_proizv = []
Ostatochn_st_oborud = []
Pogashenie_dolg_kred=[]
for i in range(len(index)):
    Obyom_proizv.append(sheet_val[index[i]+'12'].value)
    Ostatochn_st_oborud.append(sheet_val[index[i]+'13'].value)
    Pogashenie_dolg_kred.append(sheet_val[index[i]+'14'].value)
#Доходы и расходы
Cost_per_unit = sheet_val['B18'].value
Sebe_stoim = sheet_val['B19'].value
S_and_M  = sheet_val['B21'].value
ZP = sheet_val['B22'].value
Others = sheet_val['B23'].value
Komm_rasx = sheet_val['B24'].value*(10**6)
Uprav_rasx = sheet_val['B25'].value*(10**6)
Summa_rasxodov = (Komm_rasx+Uprav_rasx)*1000

#Период оборота (в днях)
S_and_M_zapasy = sheet_val['B27'].value
Nezav_proizv=sheet_val['B28'].value
Zapasi_got_prod = sheet_val['B29'].value
Debet = sheet_val['B30'].value
Kredit = sheet_val['B31'].value
Day_per_year=sheet_val['B32'].value

#Подсчет количества оборотов в год
Number_of_part = []
for i in range(5):
    Number_of_part.append(Day_per_year/(sheet_val['B'+str(27+i)].value))
print("Подсчет количества оборотов в год - ", Number_of_part)

Vyruchka_po_godam = []
Vyruchka_po_godam_min_nalog=[]
Valovaya_pribyl = []
Pribyl_ot_prodazh = []
for i in range (5):
    Vyruchka_po_godam.append(Obyom_proizv[i]*Max_production*Cost_per_unit)
    Vyruchka_po_godam_min_nalog.append(Vyruchka_po_godam[i]*(1-Nalogovaya_nagruzka))
    Valovaya_pribyl.append(Vyruchka_po_godam_min_nalog[i]-(Max_production*Obyom_proizv[i]*(Sebe_stoim))-Amortization)
    Pribyl_ot_prodazh.append(Valovaya_pribyl[i]-Summa_rasxodov)
Finish_step()
